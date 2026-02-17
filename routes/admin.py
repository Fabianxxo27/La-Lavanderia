"""
Blueprint de admin
Rutas del panel de administración
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, Response, jsonify
from werkzeug.security import generate_password_hash
from models import run_query, ensure_cliente_exists
from services import limpiar_texto, validar_email, send_email_async
from decorators import login_requerido, admin_requerido
from helpers import admin_only, obtener_esquema_descuento_cliente, ejecutar_sql_file, get_safe_redirect
from io import BytesIO
import pandas as pd
import datetime
import barcode
from barcode.writer import ImageWriter
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from pyzbar.pyzbar import decode
from PIL import Image as PILImage
import cv2
import numpy as np

bp = Blueprint('admin', __name__)



# -----------------------------------------------
@bp.route('/logout')
@login_requerido
def logout():
    session.clear()


            {"q": f"%{q}%"},
            fetchall=True
        )
        total_count = count_result[0][0] if count_result else 0
        total_paginas = (total_count + por_pagina - 1) // por_pagina
        
        # Ajustar página si está fuera de rango
        if pagina < 1:
            pagina = 1
        if pagina > total_paginas and total_paginas > 0:
            pagina = total_paginas
        
        # Obtener datos con paginación
        offset = (pagina - 1) * por_pagina
        data = run_query(
            f"SELECT id_usuario, nombre, username, email FROM usuario WHERE rol = 'cliente' AND (nombre LIKE :q OR email LIKE :q OR username LIKE :q) ORDER BY id_usuario {orden_sql} LIMIT {por_pagina} OFFSET {offset}",
            {"q": f"%{q}%"},
            fetchall=True
        )
    else:
        # Contar total de clientes
        count_result = run_query(
            "SELECT COUNT(*) FROM usuario WHERE rol = 'cliente'",
            fetchall=True
        )
        total_count = count_result[0][0] if count_result else 0
        total_paginas = (total_count + por_pagina - 1) // por_pagina
        
        # Ajustar página si está fuera de rango
        if pagina < 1:
            pagina = 1
        if pagina > total_paginas and total_paginas > 0:
            pagina = total_paginas
        
        # Obtener datos con paginación
        offset = (pagina - 1) * por_pagina
        data = run_query(
            f"SELECT id_usuario, nombre, username, email FROM usuario WHERE rol = 'cliente' ORDER BY id_usuario {orden_sql} LIMIT {por_pagina} OFFSET {offset}",
            fetchall=True
        )
    
    # Calcular rango de registros mostrados
    registro_desde = offset + 1 if total_count > 0 else 0
    registro_hasta = min(offset + por_pagina, total_count)
    
    return render_template('clientes.html', 
                         clients=data, 
                         orden=orden,
                         pagina=pagina,
                         total_paginas=total_paginas,
                         total_count=total_count,
                         registro_desde=registro_desde,
                         registro_hasta=registro_hasta)




        errores = []
        
        if not nombre or len(nombre) < 3:
            errores.append("El nombre debe tener al menos 3 caracteres.")
        
        if not username or len(username) < 3:
            errores.append("El username debe tener al menos 3 caracteres.")
        
        if not validar_email(email):
            errores.append("Por favor ingresa un email válido.")
        
        # Validar contraseña
        validacion_password = validar_contrasena(password)
        if validacion_password != True:
            errores.append(validacion_password)
        
        if password != password2:
            errores.append("Las contraseñas no coinciden.")
        
        # Verificar si el username o email ya existen
        if not errores:
            usuario_existente = run_query(
                "SELECT id_usuario FROM usuario WHERE username = :u OR email = :e",
                {"u": username, "e": email},
                fetchone=True
            )
            if usuario_existente:
                errores.append("El username o email ya están registrados.")
        
        if errores:
            for error in errores:
                flash(error, 'warning')
            return render_template('agregar_cliente.html', 
                                 nombre=nombre, 
                                 username=username, 
                                 email=email)
        
        try:
            # Hashear la contraseña
            hashed_password = generate_password_hash(password)
            
            # Insertar en la tabla usuario
            run_query(
                "INSERT INTO usuario (nombre, username, email, password, rol) VALUES (:n, :u, :e, :p, 'cliente')",
                {"n": nombre, "u": username, "e": email, "p": hashed_password},
                commit=True


            flash(f'❌ Error al agregar cliente: {str(e)}', 'danger')
    
    return render_template('agregar_cliente.html')


# -----------------------------------------------
# ACTUALIZAR CLIENTE
# -----------------------------------------------
@bp.route('/actualizar_cliente/<int:id_cliente>', methods=['GET', 'POST'])
def actualizar_cliente(id_cliente):
    """Actualizar datos de un cliente."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    cliente = run_query(
        "SELECT id_cliente, nombre, email, telefono, direccion FROM cliente WHERE id_cliente = :id",
        {"id": id_cliente},
        fetchone=True
    )
    
    if not cliente:
        flash('Cliente no encontrado.', 'danger')
        return redirect(url_for('clientes'))
    
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        telefono = request.form.get('telefono')


                "UPDATE cliente SET nombre = :n, email = :e, telefono = :t, direccion = :d WHERE id_cliente = :id",
                {"n": nombre, "e": email, "t": telefono, "d": direccion, "id": id_cliente},
                commit=True
            )
            flash('Cliente actualizado correctamente.', 'success')
            return redirect(_get_safe_redirect())
        except Exception as e:
            flash(f'Error al actualizar cliente: {e}', 'danger')
    
    return render_template('actualizar_cliente.html', cliente=cliente, id_cliente=id_cliente)




        GROUP BY p.id_pedido, p.fecha_ingreso, p.fecha_entrega, p.estado
        ORDER BY p.fecha_ingreso DESC
        LIMIT :limit OFFSET :offset
    """, {"id": id_usuario, "limit": por_pagina, "offset": offset}, fetchall=True)
    
    # Estadísticas del cliente
    stats = {
        'total_pedidos': total_count,
        'pendientes': run_query(
            "SELECT COUNT(*) FROM pedido WHERE id_cliente = :id AND estado = 'Pendiente'",
            {"id": id_usuario},
            fetchone=True
        )[0],
        'en_proceso': run_query(
            "SELECT COUNT(*) FROM pedido WHERE id_cliente = :id AND estado = 'En proceso'",
            {"id": id_usuario},
            fetchone=True
        )[0],
        'completados': run_query(
            "SELECT COUNT(*) FROM pedido WHERE id_cliente = :id AND estado = 'Completado'",
            {"id": id_usuario},
            fetchone=True
        )[0]
    }

    # Calcular paginación
    total_paginas = (total_count + por_pagina - 1) // por_pagina
    tiene_anterior = pagina > 1
    tiene_siguiente = pagina < total_paginas

    return render_template('cliente_pedidos.html', 
                         pedidos=pedidos, 
                         username=username,
                         stats=stats,
                         pagina=pagina,
                         total_paginas=total_paginas,
                         tiene_anterior=tiene_anterior,
                         tiene_siguiente=tiene_siguiente,
                         total_count=total_count)


# -----------------------------------------------
# LISTAR PEDIDOS (Administrador)
# -----------------------------------------------
@bp.route('/pedidos')
@login_requerido
@admin_requerido
def pedidos():
    """Mostrar todos los pedidos con búsqueda, filtrado y paginación (para administrador)."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    # Obtener parámetros de filtrado y paginación
    cliente_filter = request.args.get('cliente', '').strip()
    estado_filter = request.args.get('estado', '').strip()
    fecha_desde = request.args.get('desde', '').strip()
    fecha_hasta = request.args.get('hasta', '').strip()
    orden = request.args.get('orden', 'desc').strip().lower()  # 'asc' o 'desc'
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 10
    
    # Construir query base para contar
    count_query = """
        SELECT COUNT(*) FROM pedido p
        LEFT JOIN cliente c ON p.id_cliente = c.id_cliente
        WHERE 1=1
    """
    
    # Construir query base para datos
    query = """
        SELECT p.id_pedido, p.fecha_ingreso, p.fecha_entrega, p.estado, c.nombre, p.codigo_barras
        FROM pedido p
        LEFT JOIN cliente c ON p.id_cliente = c.id_cliente
        WHERE 1=1
    """
    params = {}
    
    # Agregar filtros dinámicamente
    filtro_where = ""
    if cliente_filter:
        filtro_where += " AND (LOWER(c.nombre) LIKE LOWER(:cliente) OR c.id_cliente = :cliente_id)"
        params['cliente'] = f"%{cliente_filter}%"
        try:
            params['cliente_id'] = int(cliente_filter)
        except:
            params['cliente_id'] = -1
    
    if estado_filter:
        filtro_where += " AND p.estado = :estado"
        params['estado'] = estado_filter
    
    if fecha_desde:
        filtro_where += " AND DATE(p.fecha_ingreso) >= :desde"
        params['desde'] = fecha_desde
    
    if fecha_hasta:
        filtro_where += " AND DATE(p.fecha_ingreso) <= :hasta"
        params['hasta'] = fecha_hasta
    


    # Contar total de registros
    total_result = run_query(count_query, params, fetchall=True)
    total_count = total_result[0][0] if total_result else 0
    total_paginas = (total_count + por_pagina - 1) // por_pagina if total_count > 0 else 1
    
    # Ajustar página si está fuera de rango
    if pagina < 1:
        pagina = 1
    if pagina > total_paginas:
        pagina = total_paginas
    
    # Agregar orden
    if orden == 'asc':
        query += " ORDER BY p.id_pedido ASC"
    else:
        query += " ORDER BY p.id_pedido DESC"
    
    # Agregar paginación
    offset = (pagina - 1) * por_pagina
    query += f" LIMIT {por_pagina} OFFSET {offset}"


        {"u": username.lower()},
        fetchone=True
    )
    rol = usuario[0].strip().lower() if usuario else 'cliente'
    
    # Prendas predefinidas con precios estimados
    prendas_default = [
        {'nombre': 'Camisa', 'precio': 5000},
        {'nombre': 'Pantalón', 'precio': 6000},
        {'nombre': 'Vestido', 'precio': 8000},
        {'nombre': 'Chaqueta', 'precio': 10000},
        {'nombre': 'Saco', 'precio': 7000},
        {'nombre': 'Falda', 'precio': 5500},
        {'nombre': 'Blusa', 'precio': 4500},
        {'nombre': 'Abrigo', 'precio': 12000},
        {'nombre': 'Suéter', 'precio': 6500},
        {'nombre': 'Jeans', 'precio': 7000},
        {'nombre': 'Corbata', 'precio': 3000},
        {'nombre': 'Bufanda', 'precio': 3500},
        {'nombre': 'Sábana', 'precio': 8000},
        {'nombre': 'Edredón', 'precio': 15000},
        {'nombre': 'Cortina', 'precio': 12000}
    ]
    
    if request.method == 'POST':
        try:
            from datetime import datetime, timedelta
            
            # 1. Determinar id_cliente
            if rol == 'administrador':
                id_cliente = request.form.get('id_cliente')
            else:
                usuario_data = run_query(
                    "SELECT id_usuario FROM usuario WHERE LOWER(username) = :u",
                    {"u": username.lower()},
                    fetchone=True
                )
                id_cliente = usuario_data[0] if usuario_data else None
            
            if not id_cliente:
                flash('Error al identificar el cliente.', 'danger')
                return redirect(url_for('agregar_pedido'))
            
            # 1.1 Obtener y validar direcciones (SERVICIO A DOMICILIO)
            direccion_recogida = request.form.get('direccion_recogida', '').strip()
            direccion_entrega = request.form.get('direccion_entrega', '').strip()
            
            # Validar que las direcciones no estén vacías
            if not direccion_recogida or len(direccion_recogida) < 10:
                flash('⚠️ Debes ingresar una dirección de recogida válida (mínimo 10 caracteres).', 'warning')
                return redirect(url_for('agregar_pedido'))
            
            if not direccion_entrega or len(direccion_entrega) < 10:
                flash('⚠️ Debes ingresar una dirección de entrega válida (mínimo 10 caracteres).', 'warning')
                return redirect(url_for('agregar_pedido'))
            
            # 2. Asegurar que existe en tabla cliente
            ensure_cliente_exists(id_cliente)
            
            # 3. Obtener prendas
            tipos = request.form.getlist('tipo[]')
            cantidades = request.form.getlist('cantidad[]')
            descripciones = request.form.getlist('descripcion[]')
            
            # Debug: verificar que las listas tengan el mismo tamaño
            print(f"DEBUG: tipos={len(tipos)}, cantidades={len(cantidades)}, descripciones={len(descripciones)}")
            print(f"DEBUG: cantidades={cantidades}")
            print(f"DEBUG: descripciones={descripciones}")
            
            if not tipos or len(tipos) == 0:
                flash('Debes agregar al menos una prenda.', 'warning')
                return redirect(url_for('agregar_pedido'))
            
            # 4. Calcular fechas
            total_prendas = sum(int(c) for c in cantidades if c)
            dias_entrega = 3 if total_prendas <= 5 else (5 if total_prendas <= 15 else 7)
            fecha_ingreso = datetime.now().strftime('%Y-%m-%d')
            fecha_entrega = (datetime.now() + timedelta(days=dias_entrega)).strftime('%Y-%m-%d')
            
            # 5. Crear pedido con código de barras y direcciones
            # IMPORTANTE: Primero calcular el descuento ANTES de crear el pedido
            
            # 5.1. Calcular descuento según ESQUEMA CONGELADO del cliente
            pedidos_count = run_query(
                "SELECT COUNT(*) FROM pedido WHERE id_cliente = :id AND estado != 'Cancelado'",
                {"id": id_cliente},
                fetchone=True
            )[0] or 0
            
            # Obtener esquema de descuento del cliente (congelado o actual)
            esquema_cliente = _obtener_esquema_descuento_cliente(id_cliente)
            
            # Determinar nivel y descuento basado en el esquema del cliente
            descuento_porcentaje_aplicado = 0
            nivel_descuento_aplicado = None
            
            for nivel_config in esquema_cliente:
                min_pedidos = nivel_config.get("min", 0)
                max_pedidos = nivel_config.get("max")
                
                if pedidos_count >= min_pedidos:
                    if max_pedidos is None or pedidos_count <= max_pedidos:
                        descuento_porcentaje_aplicado = nivel_config.get("porcentaje", 0)
                        nivel_descuento_aplicado = nivel_config.get("nivel")
                        break
            
            # 5.2. Crear el pedido (compatible con BD con o sin columnas de descuento)
            try:
                # Intentar con columnas de descuento (nueva estructura)
                result = run_query(
                    """INSERT INTO pedido (fecha_ingreso, fecha_entrega, estado, id_cliente, direccion_recogida, direccion_entrega, porcentaje_descuento, nivel_descuento) 
                       VALUES (:fi, :fe, :e, :ic, :dr, :de, :pd, :nd) RETURNING id_pedido""",
                    {"fi": fecha_ingreso, "fe": fecha_entrega, "e": "Pendiente", "ic": id_cliente, "dr": direccion_recogida, "de": direccion_entrega, "pd": descuento_porcentaje_aplicado, "nd": nivel_descuento_aplicado},
                    commit=True,
                    fetchone=True
                )
            except Exception as e:
                # Si falla (columnas no existen), usar estructura antigua
                if "porcentaje_descuento" in str(e) or "does not exist" in str(e):
                    result = run_query(
                        """INSERT INTO pedido (fecha_ingreso, fecha_entrega, estado, id_cliente, direccion_recogida, direccion_entrega) 
                           VALUES (:fi, :fe, :e, :ic, :dr, :de) RETURNING id_pedido""",
                        {"fi": fecha_ingreso, "fe": fecha_entrega, "e": "Pendiente", "ic": id_cliente, "dr": direccion_recogida, "de": direccion_entrega},
                        commit=True,
                        fetchone=True
                    )
                else:
                    raise e
            
            if not result or len(result) == 0:
                flash('Error al crear el pedido.', 'danger')
                return redirect(url_for('agregar_pedido'))
            
            id_pedido = result[0]
            
            # 5.1. Generar código de barras único (formato: LAV-YYYYMMDD-000001)
            codigo_barras = f"LAV-{datetime.now().strftime('%Y%m%d')}-{id_pedido:06d}"
            
            # 5.2. Actualizar pedido con código de barras
            run_query(
                "UPDATE pedido SET codigo_barras = :cb WHERE id_pedido = :id",
                {"cb": codigo_barras, "id": id_pedido},
                commit=True
            )
            
            # 6. Procesar y calcular costo primero
            prendas_a_insertar = []
            total_costo = 0
            
            for i in range(len(tipos)):
                tipo = tipos[i]
                if not tipo or tipo.strip() == '':
                    continue
                
                # Procesar cantidad con validación robusta
                cantidad_str = cantidades[i] if i < len(cantidades) else '0'
                try:
                    cantidad = int(cantidad_str) if cantidad_str and cantidad_str.strip() else 0
                    # Si la cantidad es 0 o negativa, saltar esta prenda
                    if cantidad <= 0:
                        continue
                except (ValueError, AttributeError):
                    continue
                
                descripcion = descripciones[i] if i < len(descripciones) else ''
                
                # Buscar precio
                precio = 5000  # default
                for prenda_def in prendas_default:
                    if prenda_def['nombre'] == tipo:
                        precio = prenda_def['precio']
                        break
                
                prendas_a_insertar.append({
                    'tipo': tipo,
                    'cantidad': cantidad,
                    'descripcion': descripcion,
                    'precio': precio
                })
                
                total_costo += precio * cantidad
            
            # Validar que haya al menos una prenda
            if not prendas_a_insertar or total_costo == 0:
                flash('Debes agregar al menos una prenda con cantidad mayor a 0.', 'warning')
                return redirect(url_for('agregar_pedido'))
            
            # 7. Insertar prendas en la base de datos
            prendas_insertadas = 0
            
            for prenda in prendas_a_insertar:
                tipo = prenda['tipo']
                cantidad = prenda['cantidad']
                descripcion = prenda['descripcion']
                
                for unidad in range(cantidad):
                    run_query(
                        "INSERT INTO prenda (tipo, descripcion, observaciones, id_pedido) VALUES (:tipo, :desc, :obs, :id_ped)",
                        {"tipo": tipo, "desc": descripcion, "obs": '', "id_ped": id_pedido},
                        commit=True
                    )
                    prendas_insertadas += 1
            
            # 8. Usar el descuento ya guardado en el pedido (no recalcular)
            # El descuento ya está en descuento_porcentaje_aplicado y nivel_descuento_aplicado
            
            # Calcular monto con descuento
            monto_descuento = (total_costo * descuento_porcentaje_aplicado) / 100
            monto_final = total_costo - monto_descuento
            
            # 9. Crear recibo con descuento
            run_query(
                """INSERT INTO recibo (id_pedido, id_cliente, monto, fecha) 
                   VALUES (:ip, :ic, :m, CURRENT_TIMESTAMP)""",
                {"ip": id_pedido, "ic": id_cliente, "m": monto_final},
                commit=True
            )
            
            # 10. Obtener datos del cliente para el flash
            cliente_data = run_query(
                "SELECT nombre, email FROM cliente WHERE id_cliente = :id",
                {"id": id_cliente},
                fetchone=True
            )
            
            # Enviar correo de confirmación de pedido creado (asíncrono)
            if cliente_data:
                nombre_cliente = cliente_data[0]
                email_cliente = cliente_data[1]
                
                html_pedido = f"""
                <html>
                    <body style="font-family: Arial, sans-serif; color: #333; max-width: 600px; margin: 0 auto;">
                        <div style="background: linear-gradient(135deg, #a6cc48 0%, #8fb933 100%); padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                            <h1 style="color: white; margin: 0;">✅ Pedido Creado</h1>
                        </div>
                        <div style="padding: 30px; background: #f9f9f9;">
                            <h2 style="color: #1a4e7b;">Hola {nombre_cliente},</h2>
                            <p style="font-size: 16px;">Tu pedido ha sido creado exitosamente.</p>
                            
                            <div style="background: white; border-left: 4px solid #a6cc48; padding: 20px; margin: 20px 0; border-radius: 5px;">
                                <h3 style="color: #1a4e7b; margin-top: 0;">📦 Información del Pedido</h3>
                                <p><strong>Número:</strong> #{id_pedido}</p>
                                <p><strong>Código:</strong> {codigo_barras}</p>
                                <p><strong>Fecha de recogida:</strong> {fecha_ingreso}</p>
                                <p><strong>Fecha estimada de entrega:</strong> {fecha_entrega}</p>
                                <p><strong>Estado:</strong> Pendiente</p>
                            </div>
                            
                            <div style="background: white; border-left: 4px solid #2196F3; padding: 20px; margin: 20px 0; border-radius: 5px;">
                                <h3 style="color: #1a4e7b; margin-top: 0;">🏠 Direcciones</h3>
                                <p><strong>Recogida:</strong> {direccion_recogida}</p>
                                <p><strong>Entrega:</strong> {direccion_entrega}</p>
                            </div>
                            
                            <div style="background: white; border-left: 4px solid #4CAF50; padding: 20px; margin: 20px 0; border-radius: 5px;">
                                <h3 style="color: #1a4e7b; margin-top: 0;">👕 Prendas: {prendas_insertadas}</h3>
                                <p><strong>Subtotal:</strong> ${total_costo:,.0f}</p>
                                {"<p><strong>Descuento (" + str(descuento_porcentaje_aplicado) + "%):</strong> -${:,.0f}</p>".format(monto_descuento) if descuento_porcentaje_aplicado > 0 else ""}
                                <p style="font-size: 18px; color: #a6cc48;"><strong>Total:</strong> ${monto_final:,.0f}</p>
                            </div>
                            
                            <div style="background: #e8f4f8; padding: 20px; margin: 20px 0; border-radius: 5px; text-align: center;">
                                <p><strong>🚚 Servicio a Domicilio</strong></p>
                                <p>Recogeremos tu ropa en la dirección indicada y la entregaremos limpia en la fecha estimada.</p>
                            </div>
                        </div>
                        <div style="background: #1a4e7b; color: white; padding: 20px; text-align: center; border-radius: 0 0 10px 10px;">
                            <p style="margin: 0;">La Lavandería - Servicio a Domicilio</p>
                        </div>
                    </body>
                </html>
                """
                send_email_async(email_cliente, f"Pedido {id_pedido} Creado - La Lavanderia", html_pedido)
            
            # Mensaje con descuento aplicado y código de barras
            if descuento_porcentaje_aplicado > 0:
                msg_descuento = f" | Nivel {nivel_descuento_aplicado}: Descuento {descuento_porcentaje_aplicado}% (-${monto_descuento:,.0f})"
            else:
                msg_descuento = ""
            
            flash(f'¡Pedido #{id_pedido} creado! Código: {codigo_barras} | {prendas_insertadas} prendas. Subtotal: ${total_costo:,.0f}{msg_descuento}. Total: ${monto_final:,.0f}. Entrega: {fecha_entrega}', 'success')
            
            # Redirigir según el rol
            if rol == 'administrador':
                return redirect(url_for('pedidos'))
            else:
                return redirect(url_for('cliente_pedidos'))
                
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('agregar_pedido'))
    
    clientes = run_query(
        "SELECT id_usuario, nombre FROM usuario WHERE rol = 'cliente' ORDER BY nombre",
        fetchall=True
    )
    
    # Obtener información de descuento del cliente actual si no es admin
    descuento_info = None
    if rol != 'administrador':
        usuario_data = run_query(
            "SELECT id_usuario FROM usuario WHERE LOWER(username) = :u",
            {"u": username.lower()},
            fetchone=True
        )
        if usuario_data:
            id_cliente = usuario_data[0]
            pedidos_count = run_query(
                "SELECT COUNT(*) FROM pedido WHERE id_cliente = :id AND estado != 'Cancelado'",
                {"id": id_cliente},
                fetchone=True
            )[0] or 0
            
            pedidos_en_ciclo = pedidos_count % 10
            if pedidos_en_ciclo == 0 and pedidos_count > 0:
                descuento_info = {"nivel": "Oro", "porcentaje": 15, "pedidos": pedidos_count}
            elif pedidos_en_ciclo >= 6:
                descuento_info = {"nivel": "Plata", "porcentaje": 10, "pedidos": pedidos_count}
            elif pedidos_en_ciclo >= 3:
                descuento_info = {"nivel": "Bronce", "porcentaje": 5, "pedidos": pedidos_count}
            else:
                descuento_info = {"nivel": "Sin nivel", "porcentaje": 0, "pedidos": pedidos_count}
    
    return render_template('agregar_pedido.html', 
                         clientes=clientes, 
                         rol=rol,
                         prendas_default=prendas_default,
                         descuento_info=descuento_info)


# -----------------------------------------------
# DETALLES DE PEDIDO
# -----------------------------------------------
@bp.route('/pedido_detalles/<int:id_pedido>')
def pedido_detalles(id_pedido):
    """Ver detalles de un pedido."""
    pedido = run_query(
        "SELECT id_pedido, fecha_ingreso, fecha_entrega, estado, id_cliente, codigo_barras, direccion_recogida, direccion_entrega FROM pedido WHERE id_pedido = :id",
        {"id": id_pedido},
        fetchone=True
    )
    
    if not pedido:
        flash('Pedido no encontrado.', 'danger')
        return redirect(url_for('pedidos'))
    
    prendas = run_query(
        "SELECT id_prenda, tipo, descripcion, observaciones FROM prenda WHERE id_pedido = :id",
        {"id": id_pedido},
        fetchall=True
    )
    
    # Obtener la página de origen para el botón regresar
    referer = request.args.get('ref') or request.referrer or ''
    return_url = referer if referer and referer != request.url else None
    
    return render_template('pedido_detalles.html', pedido=pedido, prendas=prendas, return_url=return_url)


# -----------------------------------------------
# ACTUALIZAR PEDIDO
# -----------------------------------------------
@bp.route('/actualizar_pedido/<int:id_pedido>', methods=['POST'])
def actualizar_pedido(id_pedido):
    """Actualizar estado de un pedido."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    estado = request.form.get('estado')
    
    try:
        # Obtener datos del pedido y cliente antes de actualizar
        pedido_data = run_query("""
            SELECT p.id_pedido, p.codigo_barras, p.fecha_entrega, c.nombre, c.email, p.id_cliente, p.estado
            FROM pedido p
            LEFT JOIN cliente c ON p.id_cliente = c.id_cliente
            WHERE p.id_pedido = :id
        """, {"id": id_pedido}, fetchone=True)
        
        if not pedido_data:
            flash('Pedido no encontrado.', 'danger')


        codigo = pedido_data[1] or f"#{id_pedido}"
        nombre_cliente = pedido_data[3]
        
        # Actualizar estado
        run_query(
            "UPDATE pedido SET estado = :e WHERE id_pedido = :id",
            {"e": estado, "id": id_pedido},
            commit=True
        )
        
        # Crear notificación para el cliente si el estado cambió
        if id_cliente and estado != estado_anterior:
            titulo = ""
            mensaje = ""
            tipo = "info"
            
            if estado == "En proceso":
                titulo = f"🔄 Pedido {codigo} en Proceso"
                mensaje = "Tu pedido está siendo procesado. Estamos lavando tu ropa con el mayor cuidado."
                tipo = "info"
            elif estado == "Completado":
                titulo = f"✅ Pedido {codigo} Completado"


                mensaje = "Tu pedido ha sido cancelado. Si tienes dudas, contacta con nosotros."
                tipo = "error"
            elif estado == "Pendiente":
                titulo = f"🕐 Pedido {codigo} Pendiente"
                mensaje = "Tu pedido está registrado y pronto será procesado."
                tipo = "warning"
            
            if titulo:
                crear_notificacion(
                    id_usuario=id_cliente,
                    titulo=titulo,
                    mensaje=mensaje,
                    tipo=tipo,
                    url=f'/cliente_pedidos'
                )
        
        # Enviar correo según el nuevo estado
        if pedido_data and pedido_data[4]:  # Si tiene email
            fecha_entrega = pedido_data[2]
            email_cliente = pedido_data[4]
            
            if estado == "En proceso":
                html = f"""
                <html>
                    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                        <div style="background: linear-gradient(135deg, #2196F3 0%, #1976D2 100%); padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                            <h1 style="color: white; margin: 0;">🔄 Pedido en Proceso</h1>
                        </div>
                        <div style="padding: 30px; background: #f9f9f9;">
                            <h2 style="color: #1a4e7b;">Hola {nombre_cliente},</h2>
                            <p>Tu pedido <strong>{codigo}</strong> está siendo procesado.</p>
                            <p>Estamos lavando tu ropa con el mayor cuidado. La entregaremos el <strong>{fecha_entrega}</strong>.</p>
                        </div>
                    </body>
                </html>
                """
                send_email_async(email_cliente, f"Pedido {codigo} en Proceso", html)
            
            elif estado == "Completado":
                html = f"""
                <html>
                    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                        <div style="background: linear-gradient(135deg, #4CAF50 0%, #388E3C 100%); padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                            <h1 style="color: white; margin: 0;">✅ Pedido Completado</h1>
                        </div>
                        <div style="padding: 30px; background: #f9f9f9;">
                            <h2 style="color: #1a4e7b;">Hola {nombre_cliente},</h2>
                            <p>¡Buenas noticias! Tu pedido <strong>{codigo}</strong> está listo.</p>
                            <p>Tu ropa está limpia y lista para ser entregada. La recibirás pronto en tu domicilio.</p>
                        </div>
                    </body>
                </html>
                """
                send_email_async(email_cliente, f"Pedido {codigo} Completado y Listo", html)
        
        flash('Pedido actualizado correctamente.', 'success')
    except Exception as e:
        flash(f'Error al actualizar: {e}', 'danger')
    
    return redirect(url_for('pedido_detalles', id_pedido=id_pedido))


# -----------------------------------------------
# ELIMINAR PEDIDO
# -----------------------------------------------
@bp.route('/eliminar_pedido/<int:id_pedido>', methods=['POST'])
def eliminar_pedido(id_pedido):
    """Eliminar un pedido y sus datos asociados (recibos y prendas)."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    try:
        # 1. Eliminar recibos asociados al pedido
        run_query(
            "DELETE FROM recibo WHERE id_pedido = :id",
            {"id": id_pedido},
            commit=True


            "DELETE FROM prenda WHERE id_pedido = :id",
            {"id": id_pedido},
            commit=True
        )
        
        # 3. Eliminar el pedido
        run_query(
            "DELETE FROM pedido WHERE id_pedido = :id",
            {"id": id_pedido},
            commit=True
        )
        flash('Pedido eliminado correctamente.', 'success')
    except Exception as e:
        flash(f'Error al eliminar: {e}', 'danger')
    
    return redirect(_get_safe_redirect())


# -----------------------------------------------
# VER PRENDAS DEL PEDIDO (CLIENTE Y ADMIN)


@login_requerido
@admin_requerido
def registro_rapido():
    """Registro rápido de cliente desde búsqueda."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    try:
        nombre = request.form.get('nombre', '').strip()
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        
        # Generar contraseña automática
        import secrets
        import string
        alphabet = string.ascii_letters + string.digits
        password_auto = ''.join(secrets.choice(alphabet) for _ in range(12))
        
        # Validación
        errores = []
        
        if not nombre or len(nombre) < 3:
            errores.append("El nombre debe tener al menos 3 caracteres.")
        
        if not username or len(username) < 3:
            errores.append("El username debe tener al menos 3 caracteres.")
        
        if not validar_email(email):
            errores.append("Email inválido.")
        
        # Verificar si el username o email ya existen
        if not errores:
            usuario_existente = run_query(
                "SELECT id_usuario FROM usuario WHERE username = :u OR email = :e",
                {"u": username, "e": email},
                fetchone=True
            )
            if usuario_existente:
                errores.append("El username o email ya están registrados.")
        
        if errores:
            for error in errores:
                flash(error, 'danger')
            return redirect(url_for('clientes'))
        
        # Hash de la contraseña generada
        hashed = generate_password_hash(password_auto, method='pbkdf2:sha256')
        
        # Insertar en usuario
        usuario_id = run_query(
            "INSERT INTO usuario (nombre, username, email, password, rol) VALUES (:n, :u, :e, :p, 'cliente') RETURNING id_usuario",
            {"n": nombre, "u": username, "e": email, "p": hashed},
            fetchone=True,
            commit=True
        )[0]
        
        # Insertar en cliente
        run_query(
            "INSERT INTO cliente (id_cliente, nombre, email) VALUES (:uid, :n, :e)",
            {"uid": usuario_id, "n": nombre, "e": email},
            commit=True
        )
        
        # Enviar email con contraseña
        html = f"""
        <html>
            <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <div style="background: linear-gradient(135deg, #a6cc48 0%, #84af1d 100%); padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                    <h1 style="color: white; margin: 0;">¡Bienvenido a La Lavandería!</h1>
                </div>
                <div style="padding: 30px; background: #f9f9f9;">


def eliminar_cliente(id_cliente):
    """Eliminar un cliente."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))


            {"id": id_cliente},
            commit=True
        )
        flash('Cliente eliminado correctamente.', 'success')
    except Exception as e:
        flash(f'Error al eliminar cliente: {e}', 'danger')
    
    return redirect(_get_safe_redirect())


# -----------------------------------------------
# TÉRMINOS Y CONDICIONES DE DESCUENTOS
# -----------------------------------------------
@bp.route('/terminos-descuentos')
def terminos_descuentos():
    """Página de términos y condiciones del programa de descuentos."""
    fecha_actual = datetime.datetime.now()
    return render_template('terminos_descuentos.html', fecha_actual=fecha_actual)



# ADMINISTRACIÓN DE DESCUENTOS
# -----------------------------------------------
@bp.route('/admin/configurar-descuentos')
@login_requerido
@admin_requerido
def configurar_descuentos():
    """Panel de administración de configuración de descuentos."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    tabla_existe = _tabla_descuento_existe()
    if not tabla_existe:
        flash('La tabla de descuentos no existe. Ejecuta las migraciones para habilitar el panel.', 'warning')
        return render_template('admin_configurar_descuentos.html', descuentos=[], tabla_descuentos_existe=False)
    


        descuentos = run_query("""
            SELECT id_config, nivel, porcentaje, pedidos_minimos, pedidos_maximos, activo
            FROM descuento_config
            ORDER BY pedidos_minimos ASC
        """, fetchall=True)
    except Exception as e:
        flash(f'Error al cargar descuentos: {e}', 'danger')
        descuentos = []
    
    return render_template('admin_configurar_descuentos.html', descuentos=descuentos, tabla_descuentos_existe=tabla_existe)


@bp.route('/admin/ejecutar-migraciones', methods=['POST'])
@login_requerido
@admin_requerido
def ejecutar_migraciones_admin():
    """Ejecutar migraciones SQL desde el panel de admin (solo si es necesario)."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))

    archivos = ['add_direcciones_to_pedido.sql', 'create_descuento_config.sql', 'add_descuento_to_pedido.sql', 'create_cliente_esquema_descuento.sql']
    errores = []

    for archivo in archivos:
        ok, err = _ejecutar_sql_file(archivo)
        if not ok:
            errores.append(f"{archivo}: {err}")

    if errores:
        flash('Errores al ejecutar migraciones: ' + ' | '.join(errores), 'danger')
    else:
        flash('Migraciones ejecutadas correctamente.', 'success')

    return redirect(url_for('configurar_descuentos'))


@bp.route('/admin/descuento/crear', methods=['POST'])
@login_requerido
@admin_requerido
def crear_descuento():
    """Crear un nuevo nivel de descuento."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    if not _tabla_descuento_existe():
        flash('La tabla de descuentos no existe. Ejecuta las migraciones antes de crear niveles.', 'warning')
        return redirect(url_for('configurar_descuentos'))
    
    nivel = request.form.get('nivel', '').strip()
    porcentaje = request.form.get('porcentaje', '').strip()
    pedidos_minimos = request.form.get('pedidos_minimos', '').strip()
    pedidos_maximos = request.form.get('pedidos_maximos', '').strip()
    activo = request.form.get('activo') == 'on'
    
    # Validación
    try:
        porcentaje = float(porcentaje)
        pedidos_minimos = int(pedidos_minimos)
        pedidos_maximos = int(pedidos_maximos) if pedidos_maximos else None
        
        if porcentaje < 0 or porcentaje > 100:
            raise ValueError("El porcentaje debe estar entre 0 y 100")
        
        if pedidos_minimos < 0:
            raise ValueError("Los pedidos mínimos deben ser positivos")
        
        if pedidos_maximos and pedidos_maximos < pedidos_minimos:
            raise ValueError("Los pedidos máximos deben ser mayores o iguales a los mínimos")
        
        # Insertar en base de datos
        run_query("""
            INSERT INTO descuento_config (nivel, porcentaje, pedidos_minimos, pedidos_maximos, activo)
            VALUES (:n, :p, :min, :max, :a)


            "p": porcentaje,
            "min": pedidos_minimos,
            "max": pedidos_maximos,
            "a": activo
        }, commit=True)
        
        flash(f'Nivel de descuento "{nivel}" creado exitosamente. Se aplicará a CLIENTES NUEVOS o que completen su ciclo actual.', 'success')
    except Exception as e:
        flash(f'Error al crear descuento: {e}', 'danger')
    
    return redirect(url_for('configurar_descuentos'))


@bp.route('/admin/descuento/editar/<int:id_config>', methods=['POST'])
@login_requerido
@admin_requerido


    
    if not _tabla_descuento_existe():
        flash('La tabla de descuentos no existe. Ejecuta las migraciones antes de editar niveles.', 'warning')
        return redirect(url_for('configurar_descuentos'))
    
    nivel = request.form.get('nivel', '').strip()
    porcentaje = request.form.get('porcentaje', '').strip()
    pedidos_minimos = request.form.get('pedidos_minimos', '').strip()
    pedidos_maximos = request.form.get('pedidos_maximos', '').strip()
    activo = request.form.get('activo') == 'on'
    
    try:
        porcentaje = float(porcentaje)
        pedidos_minimos = int(pedidos_minimos)
        pedidos_maximos = int(pedidos_maximos) if pedidos_maximos else None
        
        if porcentaje < 0 or porcentaje > 100:
            raise ValueError("El porcentaje debe estar entre 0 y 100")
        
        if pedidos_minimos < 0:
            raise ValueError("Los pedidos mínimos deben ser positivos")
        
        if pedidos_maximos and pedidos_maximos < pedidos_minimos:
            raise ValueError("Los pedidos máximos deben ser mayores o iguales a los mínimos")
        
        # Actualizar en base de datos
        run_query("""
            UPDATE descuento_config
            SET nivel = :n, porcentaje = :p, pedidos_minimos = :min, 
                pedidos_maximos = :max, activo = :a, fecha_modificacion = CURRENT_TIMESTAMP
            WHERE id_config = :id
        """, {
            "n": nivel,
            "p": porcentaje,
            "min": pedidos_minimos,
            "max": pedidos_maximos,
            "a": activo,
            "id": id_config
        }, commit=True)
        
        flash(f'Nivel de descuento "{nivel}" actualizado exitosamente. Los cambios se aplicarán solo a CLIENTES NUEVOS o que completen su ciclo actual.', 'success')
    except Exception as e:
        flash(f'Error al editar descuento: {e}', 'danger')
    
    return redirect(url_for('configurar_descuentos'))


@bp.route('/admin/descuento/eliminar/<int:id_config>', methods=['POST'])
@login_requerido
@admin_requerido
def eliminar_descuento(id_config):
    """Eliminar un nivel de descuento."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    if not _tabla_descuento_existe():
        flash('La tabla de descuentos no existe. Ejecuta las migraciones antes de eliminar niveles.', 'warning')
        return redirect(url_for('configurar_descuentos'))
    
    try:
        run_query("""
            DELETE FROM descuento_config WHERE id_config = :id
        """, {"id": id_config}, commit=True)
        
        flash('Nivel de descuento eliminado exitosamente.', 'success')
    except Exception as e:
        flash(f'Error al eliminar descuento: {e}', 'danger')
    
    return redirect(url_for('configurar_descuentos'))


# -----------------------------------------------
# REPORTES
# -----------------------------------------------
@bp.route('/reportes')
@login_requerido
@admin_requerido
def reportes():
    """Página de reportes avanzados para administrador."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    import json
    from datetime import datetime, timedelta
    
    # 1. CLIENTES NUEVOS (últimos 30 días)
    clientes_nuevos = run_query("""
        SELECT p.fecha_ingreso::date as fecha, COUNT(DISTINCT p.id_cliente) as cantidad
        FROM pedido p
        WHERE p.fecha_ingreso >= CURRENT_DATE - INTERVAL '30 days'
        GROUP BY p.fecha_ingreso::date
        ORDER BY fecha
    """, fetchall=True) or []
    
    # 2. PEDIDOS POR DÍA (últimos 30 días)
    pedidos_por_dia = run_query("""
        SELECT fecha_ingreso::date as fecha, COUNT(*) as cantidad
        FROM pedido
        WHERE fecha_ingreso >= CURRENT_DATE - INTERVAL '30 days'
        GROUP BY fecha_ingreso::date
        ORDER BY fecha
    """, fetchall=True) or []
    
    # 3. TIPOS DE PRENDAS Y SU CANTIDAD
    prendas_por_tipo = run_query("""
        SELECT tipo, COUNT(*) as cantidad
        FROM prenda
        GROUP BY tipo
        ORDER BY cantidad DESC
    """, fetchall=True) or []
    
    # 4. INGRESOS POR MES
    ingresos_por_mes = run_query("""
        SELECT DATE_TRUNC('month', fecha)::date as mes, SUM(monto) as total
        FROM recibo
        GROUP BY DATE_TRUNC('month', fecha)
        ORDER BY mes DESC
        LIMIT 12
    """, fetchall=True) or []
    
    # 5. ESTADO DE PEDIDOS
    estado_pedidos = run_query("""
        SELECT estado, COUNT(*) as cantidad
        FROM pedido
        GROUP BY estado
    """, fetchall=True) or []
    
    # 6. TOP 10 CLIENTES POR CANTIDAD DE PEDIDOS
    top_clientes = run_query("""
        SELECT c.nombre, COUNT(p.id_pedido) as cantidad_pedidos
        FROM cliente c
        LEFT JOIN pedido p ON c.id_cliente = p.id_cliente
        GROUP BY c.id_cliente, c.nombre
        ORDER BY cantidad_pedidos DESC
        LIMIT 10
    """, fetchall=True) or []
    
    # 7. ESTADÍSTICAS GENERALES
    total_clientes = run_query("SELECT COUNT(*) FROM cliente", fetchone=True)[0] or 0
    total_pedidos = run_query("SELECT COUNT(*) FROM pedido", fetchone=True)[0] or 0
    total_ingresos = run_query("SELECT COALESCE(SUM(monto), 0) FROM recibo", fetchone=True)[0] or 0
    total_prendas = run_query("SELECT COUNT(*) FROM prenda", fetchone=True)[0] or 0
    
    # 8. PROMEDIO DE PRENDAS POR PEDIDO
    promedio_prendas = run_query("""
        SELECT AVG(cantidad) as promedio
        FROM (
            SELECT COUNT(*) as cantidad
            FROM prenda
            GROUP BY id_pedido
        ) subq
    """, fetchone=True)[0] or 0
    
    # ====== NUEVAS MÉTRICAS ======
    
    # 9. ESTADO DE PEDIDOS CON CONTEO
    estado_pedidos_conteo = run_query("""
        SELECT estado, COUNT(*) as cantidad
        FROM pedido
        GROUP BY estado
    """, fetchall=True) or []
    
    # 10. TOP 5 PRENDAS MÁS SOLICITADAS
    prendas_top = run_query("""
        SELECT tipo, COUNT(*) as cantidad
        FROM prenda
        GROUP BY tipo
        ORDER BY cantidad DESC
        LIMIT 5
    """, fetchall=True) or []
    
    # 11. CLIENTES MÁS ACTIVOS CON ESTADÍSTICAS
    clientes_activos = run_query("""
        SELECT 
            c.id_cliente, 
            c.nombre,
            COUNT(DISTINCT CASE WHEN p.id_pedido IS NOT NULL THEN p.id_pedido END) as cantidad_pedidos,
            COUNT(pr.id_prenda) as total_prendas,
            COALESCE(SUM(
                CASE 
                    WHEN pr.tipo IS NULL THEN 0
                    WHEN pr.tipo = 'Camisa' THEN 5000
                    WHEN pr.tipo = 'Pantalón' THEN 6000
                    WHEN pr.tipo = 'Vestido' THEN 8000
                    WHEN pr.tipo = 'Chaqueta' THEN 10000
                    WHEN pr.tipo = 'Saco' THEN 7000
                    WHEN pr.tipo = 'Falda' THEN 5500
                    WHEN pr.tipo = 'Blusa' THEN 4500
                    WHEN pr.tipo = 'Abrigo' THEN 12000
                    WHEN pr.tipo = 'Suéter' THEN 6500
                    WHEN pr.tipo = 'Jeans' THEN 7000
                    WHEN pr.tipo = 'Corbata' THEN 3000
                    WHEN pr.tipo = 'Bufanda' THEN 3500
                    WHEN pr.tipo = 'Sábana' THEN 8000
                    WHEN pr.tipo = 'Edredón' THEN 15000
                    WHEN pr.tipo = 'Cortina' THEN 12000
                    ELSE 5000
                END
            ), 0)::numeric as gasto_total
        FROM cliente c
        LEFT JOIN pedido p ON c.id_cliente = p.id_cliente
        LEFT JOIN prenda pr ON p.id_pedido = pr.id_pedido
        GROUP BY c.id_cliente, c.nombre
        ORDER BY cantidad_pedidos DESC
        LIMIT 15
    """, fetchall=True) or []
    
    # 12. TASA DE COMPLETACIÓN
    completados = run_query("""
        SELECT COUNT(*) FROM pedido WHERE estado = 'Completado'
    """, fetchone=True)[0] or 0
    tasa_completacion = (completados / total_pedidos * 100) if total_pedidos > 0 else 0
    
    # 13. PROMEDIO DE GASTO POR CLIENTE


        FROM (
            SELECT COALESCE(SUM(
                CASE 
                    WHEN pr.tipo IS NULL THEN 0
                    WHEN pr.tipo = 'Camisa' THEN 5000
                    WHEN pr.tipo = 'Pantalón' THEN 6000
                    WHEN pr.tipo = 'Vestido' THEN 8000
                    WHEN pr.tipo = 'Chaqueta' THEN 10000
                    WHEN pr.tipo = 'Saco' THEN 7000
                    WHEN pr.tipo = 'Falda' THEN 5500
                    WHEN pr.tipo = 'Blusa' THEN 4500
                    WHEN pr.tipo = 'Abrigo' THEN 12000
                    WHEN pr.tipo = 'Suéter' THEN 6500
                    WHEN pr.tipo = 'Jeans' THEN 7000
                    WHEN pr.tipo = 'Corbata' THEN 3000
                    WHEN pr.tipo = 'Bufanda' THEN 3500
                    WHEN pr.tipo = 'Sábana' THEN 8000
                    WHEN pr.tipo = 'Edredón' THEN 15000
                    WHEN pr.tipo = 'Cortina' THEN 12000
                    ELSE 5000
                END
            ), 0) as gasto
            FROM cliente c
            LEFT JOIN pedido p ON c.id_cliente = p.id_cliente
            LEFT JOIN prenda pr ON p.id_pedido = pr.id_pedido
            GROUP BY c.id_cliente
        ) subq
    """, fetchone=True)[0] or 0
    
    # 14. PEDIDOS PENDIENTES
    pedidos_pendientes = run_query("""
        SELECT COUNT(*) FROM pedido WHERE estado IN ('Pendiente', 'En proceso')
    """, fetchone=True)[0] or 0
    
    # 15. PROMEDIO DE DÍAS PARA COMPLETAR PEDIDO
    promedio_dias = run_query("""
        SELECT AVG((fecha_entrega - fecha_ingreso)::integer)
        FROM pedido
        WHERE estado = 'Completado' AND fecha_entrega IS NOT NULL
    """, fetchone=True)[0] or 0
    
    # Preparar datos para gráficos (formato JSON)
    graficos = {
        'clientes_nuevos': {
            'labels': [str(row[0]) for row in clientes_nuevos],
            'data': [row[1] for row in clientes_nuevos]
        },
        'pedidos_dia': {
            'labels': [str(row[0]) for row in pedidos_por_dia],
            'data': [row[1] for row in pedidos_por_dia]
        },
        'prendas_tipo': {
            'labels': [row[0] for row in prendas_por_tipo],
            'data': [row[1] for row in prendas_por_tipo]
        },
        'estado_pedidos': {
            'labels': [row[0] for row in estado_pedidos],
            'data': [row[1] for row in estado_pedidos]
        },
        'top_clientes': {
            'labels': [row[0] for row in top_clientes],
            'data': [row[1] for row in top_clientes]
        },
        'ingresos_mes': {
            'labels': [str(row[0]) if row[0] else 'N/A' for row in ingresos_por_mes],
            'data': [float(row[1]) if row[1] else 0 for row in ingresos_por_mes]
        }
    }
    
    return render_template('reportes.html',
                         graficos=json.dumps(graficos),
                         total_clientes=total_clientes,
                         total_pedidos=total_pedidos,
                         total_ingresos=float(total_ingresos),
                         total_prendas=total_prendas,
                         promedio_prendas=round(float(promedio_prendas), 2),
                         estado_pedidos=estado_pedidos_conteo,
                         prendas_top=prendas_top,
                         clientes_activos=clientes_activos,
                         tasa_completacion=round(tasa_completacion, 2),
                         promedio_gasto=round(float(promedio_gasto), 0),
                         pedidos_pendientes=pedidos_pendientes,
                         promedio_dias=round(float(promedio_dias), 1))


@bp.route('/reportes/export_excel')
@login_requerido
@admin_requerido
def reportes_export_excel():
    """Exportar todos los reportes a un archivo Excel."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    try:
        from io import BytesIO
        import pandas as pd
        from datetime import datetime
        
        print("\nGenerando archivo Excel...")
        
        # Crear un buffer en memoria
        output = BytesIO()

        def safe_scalar(query, default=0):
            try:
                result = run_query(query, fetchone=True)
                if not result or result[0] is None:
                    return default
                return result[0]
            except Exception as e:
                print(f"[Resumen - Error: {e}]")
                return default
        
        # Crear el archivo Excel con múltiples hojas
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # Hoja 1: Resumen General Expandido (SIEMPRE se crea)
                try:
                    total_pedidos = safe_scalar("SELECT COUNT(*) FROM pedido", default=0)
                    total_completados = safe_scalar("SELECT COUNT(*) FROM pedido WHERE estado = 'Completado'", default=0)
                    total_clientes = safe_scalar("SELECT COUNT(*) FROM cliente", default=0)
                    total_ingresos = safe_scalar("SELECT COALESCE(SUM(total), 0) FROM recibo", default=0)
                    total_prendas = safe_scalar("SELECT COUNT(*) FROM prenda", default=0)
                    
                    # Métricas avanzadas
                    ticket_promedio = total_ingresos / max(total_completados, 1)
                    valor_por_prenda = total_ingresos / max(total_prendas, 1)
                    clientes_activos_mes = safe_scalar("""
                        SELECT COUNT(DISTINCT id_cliente) FROM pedido 
                        WHERE fecha_ingreso >= CURRENT_DATE - INTERVAL '30 days'
                    """, default=0)
                    clientes_recurrentes = safe_scalar("""
                        SELECT COUNT(*) FROM (
                            SELECT id_cliente FROM pedido GROUP BY id_cliente HAVING COUNT(*) > 1
                        ) subq
                    """, default=0)
                    
                    resumen_data = {
                        'Metrica': [
                            'Total Clientes Registrados',
                            'Clientes Activos (Ultimos 30 dias)',
                            'Clientes Recurrentes',
                            'Tasa de Retencion (%)',
                            'Total Pedidos',
                            'Pedidos Completados',
                            'Pedidos Pendientes',
                            'Pedidos En Proceso',
                            'Tasa Completacion (%)',
                            'Total Prendas Procesadas',
                            'Promedio Prendas por Pedido',
                            'Total Ingresos (COP)',
                            'Ticket Promedio (COP)',
                            'Valor Promedio por Prenda (COP)',
                            'Ingreso Promedio por Cliente (COP)',
                            'Promedio Dias para Completar',
                            'Pedidos por Dia (Promedio)',
                            'Tasa de Crecimiento Mensual (%)'
                        ],
                        'Valor': [
                            total_clientes,
                            clientes_activos_mes,
                            clientes_recurrentes,
                            round((clientes_recurrentes / max(total_clientes, 1)) * 100, 2),
                            total_pedidos,
                            total_completados,
                            safe_scalar("SELECT COUNT(*) FROM pedido WHERE estado = 'Pendiente'", default=0),
                            safe_scalar("SELECT COUNT(*) FROM pedido WHERE estado = 'En proceso'", default=0),
                            round((total_completados / max(total_pedidos, 1)) * 100, 2),
                            total_prendas,
                            round(safe_scalar("SELECT AVG(cnt) FROM (SELECT COUNT(*) as cnt FROM prenda GROUP BY id_pedido) subq", default=0), 2),
                            round(total_ingresos, 2),
                            round(ticket_promedio, 2),
                            round(valor_por_prenda, 2),
                            round(total_ingresos / max(total_clientes, 1), 2),
                            round(safe_scalar("SELECT AVG((fecha_entrega - fecha_ingreso)::integer) FROM pedido WHERE estado = 'Completado' AND fecha_entrega IS NOT NULL", default=0), 1),
                            round(safe_scalar("SELECT COUNT(*)::float / NULLIF(COUNT(DISTINCT fecha_ingreso::date), 0) FROM pedido", default=0), 2),
                            round(safe_scalar("""
                                SELECT CASE 
                                    WHEN mes_anterior > 0 THEN ((mes_actual - mes_anterior)::float / mes_anterior) * 100
                                    ELSE 0 
                                END
                                FROM (
                                    SELECT 
                                        COUNT(*) FILTER (WHERE fecha_ingreso >= DATE_TRUNC('month', CURRENT_DATE)) as mes_actual,
                                        COUNT(*) FILTER (WHERE fecha_ingreso >= DATE_TRUNC('month', CURRENT_DATE - INTERVAL '1 month') 
                                                         AND fecha_ingreso < DATE_TRUNC('month', CURRENT_DATE)) as mes_anterior
                                    FROM pedido
                                ) subq
                            """, default=0), 2)
                        ]
                    }
                    df_resumen = pd.DataFrame(resumen_data)
                except Exception as e:
                    df_resumen = pd.DataFrame({'Metrica': ['Error'], 'Valor': [f'No se pudo generar resumen: {e}']})
                df_resumen.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False)
                print("[Resumen Ejecutivo]")
            
                # Hoja 2: Estado de Pedidos con porcentajes
                try:
                    estado_data = run_query("""
                        SELECT estado, COUNT(*) as cantidad
                        FROM pedido
                        GROUP BY estado
                        ORDER BY cantidad DESC
                    """, fetchall=True)
                    if estado_data and len(estado_data) > 0:
                        df_estado = pd.DataFrame(estado_data, columns=['Estado', 'Cantidad'])
                        df_estado['Porcentaje'] = (df_estado['Cantidad'] / df_estado['Cantidad'].sum() * 100).round(2)
                        df_estado.to_excel(writer, sheet_name='Estados', index=False)
                        print("[Estados]")
                except Exception as e:
                    print(f"[Estados - Error: {e}]")
                
                # Hoja 3: Prendas por Tipo (todas)
                try:
                    prendas_data = run_query("""
                        SELECT tipo, COUNT(*) as cantidad
                        FROM prenda
                        GROUP BY tipo
                        ORDER BY cantidad DESC
                    """, fetchall=True)
                    if prendas_data and len(prendas_data) > 0:
                        df_prendas = pd.DataFrame(prendas_data, columns=['Tipo Prenda', 'Cantidad'])
                        total_prendas_tipo = df_prendas['Cantidad'].sum()
                        df_prendas['Porcentaje'] = (df_prendas['Cantidad'] / total_prendas_tipo * 100).round(2)
                        df_prendas.to_excel(writer, sheet_name='Prendas por Tipo', index=False)
                        print("[Prendas por Tipo]")
                except Exception as e:
                    print(f"[Prendas - Error: {e}]")
                
                # Hoja 4: Top 50 Clientes con Analisis Detallado
                try:
                    clientes_data = run_query("""
                        SELECT 
                            c.nombre,
                            c.email,
                            c.telefono,
                            COUNT(DISTINCT p.id_pedido) as total_pedidos,
                            COUNT(DISTINCT CASE WHEN p.estado = 'Completado' THEN p.id_pedido END) as pedidos_completados,
                            COUNT(DISTINCT CASE WHEN p.estado IN ('Pendiente', 'En proceso') THEN p.id_pedido END) as pedidos_activos,
                            COUNT(pr.id_prenda) as total_prendas,
                            COALESCE(SUM(r.total), 0) as total_gastado,
                            COALESCE(AVG(r.total), 0) as ticket_promedio,
                            MAX(p.fecha_ingreso)::date as ultimo_pedido,
                            MIN(p.fecha_ingreso)::date as primer_pedido,
                            EXTRACT(days FROM (MAX(p.fecha_ingreso) - MIN(p.fecha_ingreso))) / NULLIF(COUNT(DISTINCT p.id_pedido) - 1, 0) as dias_entre_pedidos
                        FROM cliente c
                        LEFT JOIN pedido p ON c.id_cliente = p.id_cliente
                        LEFT JOIN prenda pr ON p.id_pedido = pr.id_pedido
                        LEFT JOIN recibo r ON p.id_pedido = r.id_pedido
                        GROUP BY c.id_cliente, c.nombre, c.email, c.telefono
                        HAVING COUNT(DISTINCT p.id_pedido) > 0
                        ORDER BY total_gastado DESC, total_pedidos DESC
                        LIMIT 50
                    """, fetchall=True)
                    if clientes_data and len(clientes_data) > 0:
                        df_clientes = pd.DataFrame(clientes_data, columns=[
                            'Nombre', 'Email', 'Telefono', 'Total Pedidos', 'Completados', 'Activos',
                            'Total Prendas', 'Total Gastado (COP)', 'Ticket Promedio (COP)', 
                            'Ultimo Pedido', 'Primer Pedido', 'Dias Entre Pedidos'
                        ])
                        df_clientes['Total Gastado (COP)'] = df_clientes['Total Gastado (COP)'].round(2)
                        df_clientes['Ticket Promedio (COP)'] = df_clientes['Ticket Promedio (COP)'].round(2)
                        df_clientes['Dias Entre Pedidos'] = df_clientes['Dias Entre Pedidos'].round(1)
                        df_clientes.to_excel(writer, sheet_name='Top 50 Clientes', index=False)
                        print("[Top 50 Clientes]")
                except Exception as e:
                    print(f"[Clientes - Error: {e}]")
                
                # Hoja 5: Pedidos por Dia (ultimos 30 dias)
                try:
                    pedidos_dia_data = run_query("""
                        SELECT fecha_ingreso::date as fecha, COUNT(*) as cantidad
                        FROM pedido
                        WHERE fecha_ingreso >= CURRENT_DATE - INTERVAL '30 days'
                        GROUP BY fecha_ingreso::date
                        ORDER BY fecha DESC
                    """, fetchall=True)
                    if pedidos_dia_data and len(pedidos_dia_data) > 0:
                        df_pedidos_dia = pd.DataFrame(pedidos_dia_data, columns=['Fecha', 'Cantidad Pedidos'])
                        df_pedidos_dia.to_excel(writer, sheet_name='Pedidos por Dia', index=False)
                        print("[Pedidos por Dia]")
                except Exception as e:
                    print(f"[Pedidos por Dia - Error: {e}]")
                
                # Hoja 6: Clientes Nuevos por Dia (ultimos 30 dias)
                try:
                    clientes_nuevos_data = run_query("""
                        SELECT p.fecha_ingreso::date as fecha, COUNT(DISTINCT p.id_cliente) as clientes
                        FROM pedido p
                        WHERE p.fecha_ingreso >= CURRENT_DATE - INTERVAL '30 days'
                        GROUP BY p.fecha_ingreso::date
                        ORDER BY fecha DESC
                    """, fetchall=True)
                    if clientes_nuevos_data and len(clientes_nuevos_data) > 0:
                        df_clientes_nuevos = pd.DataFrame(clientes_nuevos_data, columns=['Fecha', 'Clientes Nuevos'])
                        df_clientes_nuevos.to_excel(writer, sheet_name='Clientes Nuevos', index=False)
                        print("[Clientes Nuevos]")
                except Exception as e:
                    print(f"[Clientes Nuevos - Error: {e}]")
                
                # Hoja 7: Ingresos por Mes (ultimos 12 meses)
                try:
                    ingresos_mes_data = run_query("""
                        SELECT 
                            TO_CHAR(fecha, 'YYYY-MM') as mes,
                            SUM(monto) as total
                        FROM recibo
                        WHERE fecha >= CURRENT_DATE - INTERVAL '12 months'
                        GROUP BY TO_CHAR(fecha, 'YYYY-MM')
                        ORDER BY mes DESC
                    """, fetchall=True)
                    if ingresos_mes_data and len(ingresos_mes_data) > 0:
                        df_ingresos_mes = pd.DataFrame(ingresos_mes_data, columns=['Mes', 'Ingresos (COP)'])
                        df_ingresos_mes.to_excel(writer, sheet_name='Ingresos por Mes', index=False)
                        print("[Ingresos por Mes]")
                except Exception as e:
                    print(f"[Ingresos por Mes - Error: {e}]")
                
                # Hoja 8: Detalle Completo de Pedidos con Rentabilidad
                try:
                    pedidos_detalle_data = run_query("""
                        SELECT 
                            p.id_pedido,
                            c.nombre as cliente,
                            c.telefono,
                            p.fecha_ingreso::date as fecha_ingreso,
                            p.fecha_entrega::date as fecha_entrega,
                            CASE 
                                WHEN p.fecha_entrega IS NOT NULL AND p.estado = 'Completado'
                                THEN (p.fecha_entrega::date - p.fecha_ingreso::date)
                                ELSE NULL
                            END as dias_procesamiento,
                            p.estado,
                            p.direccion_entrega,
                            p.direccion_recogida,
                            COUNT(pr.id_prenda) as cantidad_prendas,
                            COALESCE(r.subtotal, 0) as subtotal,
                            COALESCE(r.descuento, 0) as descuento,
                            COALESCE(r.total, 0) as total,
                            r.metodo_pago,
                            p.observaciones
                        FROM pedido p
                        LEFT JOIN cliente c ON p.id_cliente = c.id_cliente
                        LEFT JOIN prenda pr ON p.id_pedido = pr.id_pedido
                        LEFT JOIN recibo r ON p.id_pedido = r.id_pedido
                        GROUP BY p.id_pedido, c.nombre, c.telefono, p.fecha_ingreso, p.fecha_entrega, 
                                 p.estado, p.direccion_entrega, p.direccion_recogida, 
                                 r.subtotal, r.descuento, r.total, r.metodo_pago, p.observaciones
                        ORDER BY p.fecha_ingreso DESC
                    """, fetchall=True)
                    if pedidos_detalle_data and len(pedidos_detalle_data) > 0:
                        df_pedidos_detalle = pd.DataFrame(pedidos_detalle_data, 
                            columns=['ID Pedido', 'Cliente', 'Telefono', 'Fecha Ingreso', 'Fecha Entrega', 
                                   'Dias Proc.', 'Estado', 'Dir. Entrega', 'Dir. Recogida', 'Cant. Prendas', 
                                   'Subtotal (COP)', 'Descuento (COP)', 'Total (COP)', 'Metodo Pago', 'Observaciones'])
                        df_pedidos_detalle.to_excel(writer, sheet_name='Detalle Pedidos', index=False)
                        print("[Detalle Pedidos]")
                except Exception as e:
                    print(f"[Detalle Pedidos - Error: {e}]")
                
                # Hoja 9: Todos los Clientes con estadisticas
                try:
                    todos_clientes_data = run_query("""
                        SELECT 
                            c.id_cliente,
                            c.nombre,
                            c.email,
                            c.telefono,
                            COUNT(DISTINCT p.id_pedido) as total_pedidos,
                            COUNT(DISTINCT CASE WHEN p.estado = 'Completado' THEN p.id_pedido END) as pedidos_completados,
                            COUNT(pr.id_prenda) as total_prendas,
                            COALESCE(SUM(r.total), 0) as total_gastado
                        FROM cliente c
                        LEFT JOIN pedido p ON c.id_cliente = p.id_cliente
                        LEFT JOIN prenda pr ON p.id_pedido = pr.id_pedido
                        LEFT JOIN recibo r ON p.id_pedido = r.id_pedido
                        GROUP BY c.id_cliente, c.nombre, c.email, c.telefono
                        ORDER BY total_pedidos DESC, total_gastado DESC
                    """, fetchall=True)
                    if todos_clientes_data and len(todos_clientes_data) > 0:
                        df_todos_clientes = pd.DataFrame(todos_clientes_data, 
                            columns=['ID', 'Nombre', 'Email', 'Telefono', 'Total Pedidos', 'Completados', 'Total Prendas', 'Total Gastado (COP)'])
                        df_todos_clientes.to_excel(writer, sheet_name='Todos los Clientes', index=False)
                        print("[Todos los Clientes]")
                except Exception as e:
                    print(f"[Todos Clientes - Error: {e}]")
                
                # Hoja 10: Prendas por Pedido (analisis detallado)
                try:
                    prendas_pedido_data = run_query("""
                        SELECT 
                            p.id_pedido,
                            c.nombre as cliente,
                            pr.tipo,
                            pr.color,
                            pr.estado as estado_prenda,
                            pr.observaciones
                        FROM prenda pr
                        JOIN pedido p ON pr.id_pedido = p.id_pedido
                        JOIN cliente c ON p.id_cliente = c.id_cliente
                        ORDER BY p.id_pedido DESC, pr.tipo
                    """, fetchall=True)
                    if prendas_pedido_data and len(prendas_pedido_data) > 0:
                        df_prendas_pedido = pd.DataFrame(prendas_pedido_data,
                            columns=['ID Pedido', 'Cliente', 'Tipo Prenda', 'Color', 'Estado', 'Observaciones'])
                        df_prendas_pedido.to_excel(writer, sheet_name='Prendas Detalle', index=False)
                        print("[Prendas Detalle]")
                except Exception as e:
                    print(f"[Prendas Detalle - Error: {e}]")
                
                # Hoja 11: Analisis de Rentabilidad Mensual
                try:
                    rentabilidad_data = run_query("""
                        SELECT 
                            TO_CHAR(r.fecha, 'YYYY-MM') as mes,
                            COUNT(DISTINCT r.id_pedido) as pedidos,
                            COUNT(DISTINCT p.id_cliente) as clientes,
                            SUM(r.subtotal) as subtotal,
                            SUM(r.descuento) as descuentos_aplicados,
                            SUM(r.total) as ingresos_netos,
                            AVG(r.total) as ticket_promedio,
                            SUM(r.subtotal) - SUM(r.total) as descuentos_totales
                        FROM recibo r
                        JOIN pedido p ON r.id_pedido = p.id_pedido
                        WHERE r.fecha >= CURRENT_DATE - INTERVAL '12 months'
                        GROUP BY TO_CHAR(r.fecha, 'YYYY-MM')
                        ORDER BY mes DESC
                    """, fetchall=True)
                    if rentabilidad_data and len(rentabilidad_data) > 0:
                        df_rentabilidad = pd.DataFrame(rentabilidad_data, columns=[
                            'Mes', 'Pedidos', 'Clientes', 'Subtotal (COP)', 'Descuentos (COP)', 
                            'Ingresos Netos (COP)', 'Ticket Promedio (COP)', 'Total Desc. (COP)'
                        ])
                        df_rentabilidad.to_excel(writer, sheet_name='Rentabilidad Mensual', index=False)
                        print("[Rentabilidad Mensual]")
                except Exception as e:
                    print(f"[Rentabilidad - Error: {e}]")
                
                # Hoja 12: Recibos Detallados
                try:
                    recibos_data = run_query("""
                        SELECT 
                            r.id_recibo,
                            r.id_pedido,
                            c.nombre as cliente,
                            r.fecha::date as fecha_emision,
                            r.subtotal,
                            r.descuento,
                            r.total,
                            r.metodo_pago,
                            r.estado as estado_pago,
                            p.estado as estado_pedido
                        FROM recibo r
                        JOIN pedido p ON r.id_pedido = p.id_pedido
                        JOIN cliente c ON p.id_cliente = c.id_cliente
                        ORDER BY r.fecha DESC
                    """, fetchall=True)
                    if recibos_data and len(recibos_data) > 0:
                        df_recibos = pd.DataFrame(recibos_data, columns=[
                            'ID Recibo', 'ID Pedido', 'Cliente', 'Fecha Emision', 
                            'Subtotal (COP)', 'Descuento (COP)', 'Total (COP)', 
                            'Metodo Pago', 'Estado Pago', 'Estado Pedido'
                        ])
                        df_recibos.to_excel(writer, sheet_name='Recibos Detallados', index=False)
                        print("[Recibos Detallados]")
                except Exception as e:
                    print(f"[Recibos - Error: {e}]")
                
                # Hoja 13: Clientes Inactivos
                try:
                    inactivos_data = run_query("""
                        SELECT 
                            c.nombre,
                            c.email,
                            c.telefono,
                            COUNT(p.id_pedido) as total_pedidos,
                            MAX(p.fecha_ingreso)::date as ultimo_pedido,
                            EXTRACT(days FROM (CURRENT_DATE - MAX(p.fecha_ingreso))) as dias_inactivo,
                            COALESCE(SUM(r.total), 0) as total_gastado
                        FROM cliente c
                        LEFT JOIN pedido p ON c.id_cliente = p.id_cliente
                        LEFT JOIN recibo r ON p.id_pedido = r.id_pedido
                        GROUP BY c.id_cliente, c.nombre, c.email, c.telefono
                        HAVING MAX(p.fecha_ingreso) < CURRENT_DATE - INTERVAL '60 days' 
                               OR MAX(p.fecha_ingreso) IS NULL
                        ORDER BY dias_inactivo DESC NULLS FIRST
                        LIMIT 100
                    """, fetchall=True)
                    if inactivos_data and len(inactivos_data) > 0:
                        df_inactivos = pd.DataFrame(inactivos_data, columns=[
                            'Nombre', 'Email', 'Telefono', 'Total Pedidos', 
                            'Ultimo Pedido', 'Dias Inactivo', 'Total Gastado (COP)'
                        ])
                        df_inactivos.to_excel(writer, sheet_name='Clientes Inactivos', index=False)
                        print("[Clientes Inactivos]")
                except Exception as e:
                    print(f"[Inactivos - Error: {e}]")
                
                # Hoja 14: Analisis de Descuentos Aplicados
                try:
                    descuentos_data = run_query("""
                        SELECT 
                            c.nombre as cliente,
                            COUNT(r.id_recibo) as pedidos_con_descuento,
                            SUM(r.descuento) as total_descuentos,
                            AVG(r.descuento) as descuento_promedio,
                            SUM(r.subtotal) as subtotal_acumulado,
                            SUM(r.total) as total_pagado,
                            (SUM(r.descuento) / NULLIF(SUM(r.subtotal), 0) * 100) as porcentaje_desc_promedio
                        FROM recibo r
                        JOIN pedido p ON r.id_pedido = p.id_pedido
                        JOIN cliente c ON p.id_cliente = c.id_cliente
                        WHERE r.descuento > 0
                        GROUP BY c.id_cliente, c.nombre
                        ORDER BY total_descuentos DESC
                        LIMIT 50
                    """, fetchall=True)
                    if descuentos_data and len(descuentos_data) > 0:
                        df_descuentos = pd.DataFrame(descuentos_data, columns=[
                            'Cliente', 'Pedidos con Descuento', 'Total Descuentos (COP)', 
                            'Descuento Promedio (COP)', 'Subtotal Acumulado (COP)', 
                            'Total Pagado (COP)', 'Porcentaje Desc. Promedio'
                        ])
                        df_descuentos['Porcentaje Desc. Promedio'] = df_descuentos['Porcentaje Desc. Promedio'].round(2)
                        df_descuentos.to_excel(writer, sheet_name='Analisis Descuentos', index=False)
                        print("[Analisis Descuentos]")
                except Exception as e:
                    print(f"[Descuentos - Error: {e}]")
                
                # Hoja 15: Prendas Mas Rentables
                try:
                    rentables_data = run_query("""
                        SELECT 
                            pr.tipo,
                            COUNT(*) as cantidad_procesada,
                            COUNT(DISTINCT pr.id_pedido) as pedidos,
                            CASE 
                                WHEN pr.tipo = 'Camisa' THEN 5000
                                WHEN pr.tipo = 'Pantalon' THEN 6000
                                WHEN pr.tipo = 'Vestido' THEN 8000
                                WHEN pr.tipo = 'Chaqueta' THEN 10000
                                WHEN pr.tipo = 'Saco' THEN 7000
                                WHEN pr.tipo = 'Falda' THEN 5500
                                WHEN pr.tipo = 'Blusa' THEN 4500
                                WHEN pr.tipo = 'Abrigo' THEN 12000
                                WHEN pr.tipo = 'Sueter' THEN 6500
                                WHEN pr.tipo = 'Jeans' THEN 7000
                                WHEN pr.tipo = 'Corbata' THEN 3000
                                WHEN pr.tipo = 'Bufanda' THEN 3500
                                WHEN pr.tipo = 'Sabana' THEN 8000
                                WHEN pr.tipo = 'Edredon' THEN 15000
                                WHEN pr.tipo = 'Cortina' THEN 12000
                                ELSE 5000
                            END as precio_unitario,
                            COUNT(*) * CASE 
                                WHEN pr.tipo = 'Camisa' THEN 5000
                                WHEN pr.tipo = 'Pantalon' THEN 6000
                                WHEN pr.tipo = 'Vestido' THEN 8000
                                WHEN pr.tipo = 'Chaqueta' THEN 10000
                                WHEN pr.tipo = 'Saco' THEN 7000
                                WHEN pr.tipo = 'Falda' THEN 5500
                                WHEN pr.tipo = 'Blusa' THEN 4500
                                WHEN pr.tipo = 'Abrigo' THEN 12000
                                WHEN pr.tipo = 'Sueter' THEN 6500
                                WHEN pr.tipo = 'Jeans' THEN 7000
                                WHEN pr.tipo = 'Corbata' THEN 3000
                                WHEN pr.tipo = 'Bufanda' THEN 3500
                                WHEN pr.tipo = 'Sabana' THEN 8000
                                WHEN pr.tipo = 'Edredon' THEN 15000
                                WHEN pr.tipo = 'Cortina' THEN 12000
                                ELSE 5000
                            END as ingreso_estimado
                        FROM prenda pr
                        GROUP BY pr.tipo
                        ORDER BY ingreso_estimado DESC
                    """, fetchall=True)
                    if rentables_data and len(rentables_data) > 0:
                        df_rentables = pd.DataFrame(rentables_data, columns=[
                            'Tipo Prenda', 'Cantidad Procesada', 'Pedidos', 
                            'Precio Unitario (COP)', 'Ingreso Estimado (COP)'
                        ])
                        df_rentables.to_excel(writer, sheet_name='Prendas Rentables', index=False)
                        print("[Prendas Rentables]")
                except Exception as e:
                    print(f"[Prendas Rentables - Error: {e}]")
                
                # Hoja 16: Rendimiento por Dia de Semana
                try:
                    semana_data = run_query("""
                        SELECT 
                            CASE EXTRACT(dow FROM fecha_ingreso)
                                WHEN 0 THEN 'Domingo'
                                WHEN 1 THEN 'Lunes'
                                WHEN 2 THEN 'Martes'
                                WHEN 3 THEN 'Miercoles'
                                WHEN 4 THEN 'Jueves'
                                WHEN 5 THEN 'Viernes'
                                WHEN 6 THEN 'Sabado'
                            END as dia_semana,
                            EXTRACT(dow FROM fecha_ingreso) as orden,
                            COUNT(*) as pedidos,
                            COUNT(DISTINCT id_cliente) as clientes_unicos,
                            COALESCE(AVG((fecha_entrega - fecha_ingreso)::integer), 0) as dias_promedio_entrega
                        FROM pedido
                        GROUP BY EXTRACT(dow FROM fecha_ingreso)
                        ORDER BY orden
                    """, fetchall=True)
                    if semana_data and len(semana_data) > 0:
                        df_semana = pd.DataFrame(semana_data, columns=[
                            'Dia Semana', 'Orden', 'Pedidos', 'Clientes Unicos', 'Dias Promedio Entrega'
                        ])
                        df_semana = df_semana.drop('Orden', axis=1)
                        df_semana.to_excel(writer, sheet_name='Rendimiento x Dia', index=False)
                        print("[Rendimiento por Dia]")
                except Exception as e:
                    print(f"[Rendimiento Dia - Error: {e}]")
        except Exception as e:
            print(f"[ERROR] export_excel fallo: {e}")
            try:
                from openpyxl import Workbook
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Resumen"
                ws["A1"] = "Error"
                ws["B1"] = f"No se pudo generar el Excel: {e}"
                wb.save(output)
                output.seek(0)
                from flask import send_file
                fecha_actual = datetime.now().strftime('%Y-%m-%d_%H-%M')
                filename = f'Reportes_LaLavanderia_{fecha_actual}.xlsx'
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
            except Exception as fallback_error:
                print(f"[ERROR] export_excel fallback fallo: {fallback_error}")
                raise
        
        print("Preparando descarga...")
        
        # IMPORTANTE: seek DESPUES de cerrar el writer
        output.seek(0)
        
        # Preparar respuesta para descarga
        from flask import send_file
        fecha_actual = datetime.now().strftime('%Y-%m-%d_%H-%M')
        filename = f'Reportes_LaLavanderia_{fecha_actual}.xlsx'
        
        return send_file(


        )
    
    except ImportError as e:
        flash(f'❌ Error: Falta instalar dependencias (pandas/openpyxl). Contacta al administrador.', 'danger')
        print(f"Error de importación en export_excel: {e}")
        return redirect(url_for('reportes'))
    except Exception as e:
        flash(f'❌ Error al generar archivo Excel: {str(e)}', 'danger')
        print(f"Error en export_excel: {e}")
        return redirect(url_for('reportes'))


# -----------------------------------------------
# LECTOR DE CÓDIGOS DE BARRAS
# -----------------------------------------------
@bp.route('/lector_barcode', methods=['GET', 'POST'])
@login_requerido
@admin_requerido
def lector_barcode():
    """Escanear código de barras y mostrar detalles del pedido."""
    if not _admin_only():
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Verificar si se subió una imagen
        if 'barcode_image' not in request.files:
            return jsonify({'success': False, 'error': 'No se subió ninguna imagen'}), 400
        
        file = request.files['barcode_image']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No se seleccionó ningún archivo'}), 400
        
        try:
            # Validar que el archivo sea una imagen
            allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
            file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
            
            if file_ext not in allowed_extensions:
                return jsonify({
                    'success': False, 
                    'error': 'Formato de archivo no válido. Por favor sube una imagen (PNG, JPG, JPEG, GIF, BMP, WEBP)'
                }), 400
            
            # Leer imagen
            image_bytes = file.read()
            
            if len(image_bytes) == 0:
                return jsonify({'success': False, 'error': 'El archivo está vacío o corrupto'}), 400
            
            # Intentar decodificar la imagen
            nparr = np.frombuffer(image_bytes, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if img is None:
                return jsonify({
                    'success': False, 
                    'error': 'No se pudo leer la imagen. Asegúrate de que el archivo sea una imagen válida'
                }), 400
            
            # Verificar que la imagen tenga un tamaño razonable
            height, width = img.shape[:2]
            if width < 50 or height < 50:
                return jsonify({
                    'success': False, 
                    'error': 'La imagen es demasiado pequeña. Por favor usa una imagen de mejor calidad'
                }), 400
            
            # Convertir a PIL Image para pyzbar
            try:
                img_pil = PILImage.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
            except Exception as conv_error:
                return jsonify({
                    'success': False, 
                    'error': 'Error al procesar la imagen. Intenta con otra imagen'
                }), 400
            
            # Decodificar código de barras
            decoded_objects = decode(img_pil)
            
            if not decoded_objects or len(decoded_objects) == 0:
                return jsonify({
                    'success': False, 
                    'error': 'No se detectó ningún código de barras en la imagen. Asegúrate de que la imagen contenga un código de barras visible y bien enfocado'
                }), 400
            
            # Obtener el primer código detectado
            try:
                barcode_data = decoded_objects[0].data.decode('utf-8')
            except UnicodeDecodeError:
                return jsonify({
                    'success': False, 
                    'error': 'El código de barras detectado no es válido o está corrupto'
                }), 400
            
            # Buscar pedido por código de barras
            try:
                pedido = run_query("""
                    SELECT p.id_pedido, p.id_cliente, p.fecha_ingreso, p.fecha_entrega, 
                           p.estado, c.nombre, c.telefono, c.direccion, u.email
                    FROM pedido p
                    JOIN cliente c ON p.id_cliente = c.id_cliente
                    LEFT JOIN usuario u ON c.id_cliente = u.id_usuario
                    WHERE p.codigo_barras = :codigo
                """, {"codigo": barcode_data}, fetchone=True)
            except Exception as db_error:
                print(f"Error en consulta de pedido: {str(db_error)}")
                return jsonify({
                    'success': False, 
                    'error': 'Error al buscar el pedido en la base de datos'
                }), 500
            
            if not pedido:
                return jsonify({'success': False, 'error': f'No se encontró ningún pedido con el código: {barcode_data}'}), 404
            
            # Obtener prendas del pedido
            try:
                prendas = run_query("""
                    SELECT tipo, descripcion, observaciones
                    FROM prenda
                    WHERE id_pedido = :id
                """, {"id": pedido[0]}, fetchall=True)
                print(f"DEBUG - Pedido ID: {pedido[0]}, Prendas encontradas: {len(prendas) if prendas else 0}")
                if prendas:
                    print(f"DEBUG - Prendas: {prendas}")
            except Exception as prendas_error:
                print(f"Error al obtener prendas: {str(prendas_error)}")
                prendas = []
            
            # Obtener recibo del pedido
            try:
                recibo = run_query("""
                    SELECT monto, descuento, fecha
                    FROM recibo
                    WHERE id_pedido = :id
                """, {"id": pedido[0]}, fetchone=True)
            except Exception as recibo_error:
                print(f"Error al obtener recibo: {str(recibo_error)}")
                recibo = None
            
            # Preparar respuesta con manejo seguro de fechas
            try:
                fecha_ingreso_str = 'N/A'
                if pedido[2]:
                    try:
                        fecha_ingreso_str = pedido[2].strftime('%d/%m/%Y %H:%M')
                    except:
                        fecha_ingreso_str = str(pedido[2])
                
                fecha_entrega_str = 'Pendiente'
                if pedido[3]:
                    try:
                        fecha_entrega_str = pedido[3].strftime('%d/%m/%Y')
                    except:
                        fecha_entrega_str = str(pedido[3])
                
                response_data = {
                    'success': True,
                    'codigo_barras': barcode_data,
                    'pedido': {
                        'id': pedido[0],
                        'fecha_ingreso': fecha_ingreso_str,
                        'fecha_entrega': fecha_entrega_str,
                        'estado': pedido[4] or 'Desconocido',
                    },
                    'cliente': {
                        'id': pedido[1],
                        'nombre': pedido[5] or 'No registrado',
                        'telefono': pedido[6] or 'No registrado',
                        'direccion': pedido[7] or 'No registrada',
                        'email': pedido[8] or 'No registrado'
                    },
                    'prendas': [{'tipo': p[0], 'descripcion': p[1] or '', 'observaciones': p[2] or ''} for p in (prendas or [])],
                    'recibo': None
                }
                
                print(f"DEBUG - Response data prendas: {response_data['prendas']}")
                
                # Agregar información del recibo si existe
                if recibo:
                    try:
                        fecha_recibo_str = 'N/A'
                        if recibo[2]:
                            try:
                                fecha_recibo_str = recibo[2].strftime('%d/%m/%Y')
                            except:
                                fecha_recibo_str = str(recibo[2])
                        
                        response_data['recibo'] = {
                            'monto': float(recibo[0]) if recibo[0] else 0,
                            'descuento': float(recibo[1]) if recibo[1] else 0,
                            'fecha': fecha_recibo_str
                        }
                    except Exception as recibo_format_error:
                        print(f"Error al formatear recibo: {str(recibo_format_error)}")
                        response_data['recibo'] = None
                
            except Exception as format_error:
                print(f"Error al formatear respuesta: {str(format_error)}")
                return jsonify({
                    'success': False, 
                    'error': 'Error al procesar los datos del pedido'
                }), 500
            
            return jsonify(response_data), 200
        
        except cv2.error as cv_error:
            return jsonify({
                'success': False, 
                'error': f'Error al procesar la imagen con OpenCV: {str(cv_error)}'
            }), 500
        
        except ValueError as val_error:
            return jsonify({
                'success': False, 
                'error': f'Datos inválidos en la imagen: {str(val_error)}'
            }), 400
        
        except MemoryError:
            return jsonify({
                'success': False, 
                'error': 'La imagen es demasiado grande. Por favor usa una imagen más pequeña'
            }), 400
            
        except Exception as e:
            # Log del error para debugging (puedes ver esto en los logs de Render)
            import traceback
            error_trace = traceback.format_exc()
            print(f"Error inesperado en lector_barcode:")
            print(error_trace)
            
            # Intentar obtener más detalles del error
            error_details = str(e)
            if 'psycopg2' in error_trace:
                error_msg = 'Error de base de datos. Por favor contacta al administrador'
            elif 'decode' in error_trace.lower():
                error_msg = 'Error al decodificar el código de barras'
            elif 'image' in error_trace.lower():
                error_msg = 'Error al procesar la imagen'
            else:
                error_msg = f'Error: {error_details[:100]}'
            
            return jsonify({
                'success': False, 
                'error': error_msg
            }), 500
    
    # GET request - mostrar página
    return render_template('lector_barcode.html')


# -----------------------------------------------
# AGREGAR PEDIDO
# -----------------------------------------------
@bp.route('/agregar_pedido', methods=['GET', 'POST'])
def agregar_pedido():
    """Crear un nuevo pedido (clientes y administradores)."""
    username = session.get('username')
    if not username:
        flash('Debes iniciar sesión para crear pedidos.', 'danger')
        return redirect(url_for('login'))
